#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Zabbix Host Import (Excel -> Zabbix API) - v7.4
COMPATÍVEL COM ZABBIX 7.4+

Principais correções para Zabbix 7.4:
- Suporte a Template Groups (separado de Host Groups)
- Uso de templategroup.get/create para templates
- Uso de hostgroup.get/create para hosts
- Template.create agora usa "groups" para template groups

Mantém:
- Auth via header Authorization: Bearer <TOKEN>
- DRY_RUN default (no --apply)
- Multiple groups/templates separated by ';'
- Create group if missing
- Create template if missing (empty container)
- Update existing host instead of duplicating
- Report CSV per row (saved beside the Excel by default)

Usage:
  python zabbix_import_excel_v7_4.py --excel "C:\\Temp\\zabbix_hosts.xlsx"
  python zabbix_import_excel_v7_4.py --excel "C:\\Temp\\zabbix_hosts.xlsx" --apply
"""

from __future__ import annotations
import sys
import json
import csv
import argparse
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from pathlib import Path
import re

import requests
from openpyxl import load_workbook

# =========================
# CONFIGURE AQUI
# =========================
ZABBIX_URL = "http://192.168.252.21/zabbix/api_jsonrpc.php"
ZABBIX_TOKEN = "54abf15176506f235f44a84a70c967b53b4e6d290e76eb2b7a07130eba1a05a8"

DEFAULT_EXCEL_NAME = "zabbix_hosts_template_v2.xlsx"
SHEET_NAME = "Hosts"

DEFAULT_TEMPLATE_GROUP = "Templates/Auto"
AGENT_PORT = "10050"
CONTINUE_ON_ERROR = True

# Se True, ao atualizar host ele vai AJUSTAR o visible name ("name") para o valor do Excel.
UPDATE_VISIBLE_NAME = True

# Se True, ao atualizar host ele vai AJUSTAR o technical name ("host") para o sanitized do Excel.
# ⚠️ Em ambientes já existentes, normalmente é melhor manter False para não renomear "host".
UPDATE_TECHNICAL_HOSTNAME = False
# =========================

HEADERS = {
    "Content-Type": "application/json-rpc",
    "Authorization": f"Bearer {ZABBIX_TOKEN}",
}
SESSION = requests.Session()


class ZabbixAPIError(RuntimeError):
    pass


def api_call(method: str, params: Dict[str, Any], request_id: int = 1) -> Dict[str, Any]:
    payload = {"jsonrpc": "2.0", "method": method, "params": params, "id": request_id}
    try:
        r = SESSION.post(ZABBIX_URL, json=payload, headers=HEADERS, timeout=60)
        r.raise_for_status()
    except Exception as e:
        raise ZabbixAPIError(f"Falha HTTP ao chamar {method}: {e}") from e

    data = r.json()
    if "error" in data:
        err = data["error"]
        raise ZabbixAPIError(f"Erro na API {method}: {err.get('message')} | {err.get('data')}")
    return data


def split_multi(value: Any) -> List[str]:
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    return [p.strip() for p in s.split(";") if p.strip()]


def sanitize_technical_name(value: str, max_len: int = 128) -> str:
    """
    Zabbix "host" (technical name) typically forbids spaces and some chars.
    Strategy:
      - trim
      - replace spaces with '_'
      - replace any char not in [A-Za-z0-9._-] with '_'
      - collapse repeated underscores
      - strip underscores from ends
      - cut to max_len
    """
    v = (value or "").strip()
    v = v.replace(" ", "_")
    v = re.sub(r"[^A-Za-z0-9\.\_\-]", "_", v)
    v = re.sub(r"_+", "_", v)
    v = v.strip("_")
    if not v:
        v = "unnamed"
    return v[:max_len]


def ensure_hostgroup(group_name: str, dry_run: bool) -> str:
    """Cria ou retorna o ID de um HOST GROUP (para hosts)"""
    res = api_call("hostgroup.get", {"filter": {"name": [group_name]}})
    if res["result"]:
        return res["result"][0]["groupid"]

    if dry_run:
        print(f"[DRY] Criaria host group: {group_name}")
        return "0"

    res2 = api_call("hostgroup.create", {"name": group_name})
    groupid = res2["result"]["groupids"][0]
    print(f"[OK] Host group criado: {group_name} (id={groupid})")
    return groupid


def ensure_templategroup(group_name: str, dry_run: bool) -> str:
    """Cria ou retorna o ID de um TEMPLATE GROUP (para templates)"""
    res = api_call("templategroup.get", {"filter": {"name": [group_name]}})
    if res["result"]:
        return res["result"][0]["groupid"]

    if dry_run:
        print(f"[DRY] Criaria template group: {group_name}")
        return "0"

    res2 = api_call("templategroup.create", {"name": group_name})
    groupid = res2["result"]["groupids"][0]
    print(f"[OK] Template group criado: {group_name} (id={groupid})")
    return groupid


def template_get_by_host_or_name(template_name: str) -> Optional[Dict[str, Any]]:
    res = api_call("template.get", {"filter": {"host": [template_name]}, "output": ["templateid", "host", "name"]})
    if res["result"]:
        return res["result"][0]

    res2 = api_call("template.get", {"filter": {"name": [template_name]}, "output": ["templateid", "host", "name"]})
    if res2["result"]:
        return res2["result"][0]

    return None


def ensure_template(template_name: str, dry_run: bool) -> str:
    existing = template_get_by_host_or_name(template_name)
    if existing:
        return existing["templateid"]

    # Para templates, usamos TEMPLATE GROUPS
    tg_id = ensure_templategroup(DEFAULT_TEMPLATE_GROUP, dry_run=dry_run)
    tech = sanitize_technical_name(template_name)

    if dry_run:
        print(f"[DRY] Criaria template VAZIO: name='{template_name}' host='{tech}' no template group '{DEFAULT_TEMPLATE_GROUP}'")
        return "0"

    # No Zabbix 7.4+, template.create usa "groups" para template groups
    res2 = api_call("template.create", {
        "host": tech,
        "name": template_name,
        "groups": [{"groupid": tg_id}],
    })
    templateid = res2["result"]["templateids"][0]
    print(f"[OK] Template criado (VAZIO): {template_name} (host={tech}, id={templateid})")
    return templateid


def host_get_by_host_or_name(value: str) -> Optional[Dict[str, Any]]:
    res = api_call("host.get", {
        "filter": {"host": [value]},
        "selectInterfaces": ["interfaceid", "type", "main", "useip", "ip", "dns", "port"],
        "selectGroups": ["groupid", "name"],
        "selectParentTemplates": ["templateid", "host", "name"],
        "output": ["hostid", "host", "name"]
    })
    if res["result"]:
        return res["result"][0]

    res2 = api_call("host.get", {
        "filter": {"name": [value]},
        "selectInterfaces": ["interfaceid", "type", "main", "useip", "ip", "dns", "port"],
        "selectGroups": ["groupid", "name"],
        "selectParentTemplates": ["templateid", "host", "name"],
        "output": ["hostid", "host", "name"]
    })
    if res2["result"]:
        return res2["result"][0]

    return None


def get_host_by_ip(ip: str) -> Optional[Dict[str, Any]]:
    res = api_call("hostinterface.get", {
        "filter": {"ip": [ip]},
        "output": ["interfaceid", "hostid", "ip", "type", "main"]
    })
    if not res["result"]:
        return None

    hostid = res["result"][0]["hostid"]
    res2 = api_call("host.get", {
        "hostids": [hostid],
        "selectInterfaces": ["interfaceid", "type", "main", "useip", "ip", "dns", "port"],
        "selectGroups": ["groupid", "name"],
        "selectParentTemplates": ["templateid", "host", "name"],
        "output": ["hostid", "host", "name"]
    })
    return res2["result"][0] if res2["result"] else None


def pick_agent_interfaceid(host: Dict[str, Any]) -> Optional[str]:
    for iface in host.get("interfaces", []):
        if str(iface.get("type")) == "1" and str(iface.get("main")) == "1":
            return iface.get("interfaceid")
    return None


def update_interface_ip(hostid: str, interfaceid: Optional[str], desired_ip: str, dry_run: bool) -> None:
    if dry_run:
        print(f"[DRY] Ajustaria interface do hostid={hostid} para IP={desired_ip}")
        return

    if interfaceid:
        api_call("hostinterface.update", {"interfaceid": interfaceid, "ip": desired_ip, "useip": 1, "dns": ""})
    else:
        api_call("hostinterface.create", {
            "hostid": hostid,
            "interfaces": [{
                "type": 1,
                "main": 1,
                "useip": 1,
                "ip": desired_ip,
                "dns": "",
                "port": AGENT_PORT
            }]
        })


def set_groups_and_templates(
    host: Dict[str, Any],
    groupids: List[str],
    templateids: List[str],
    desired_visible_name: str,
    desired_technical_host: str,
    dry_run: bool
) -> None:
    hostid = host["hostid"]

    current_tids = [t["templateid"] for t in host.get("parentTemplates", [])]
    to_clear = [tid for tid in current_tids if tid not in templateids]

    params: Dict[str, Any] = {
        "hostid": hostid,
        "groups": [{"groupid": gid} for gid in groupids],
        "templates": [{"templateid": tid} for tid in templateids],
    }
    if to_clear:
        params["templates_clear"] = [{"templateid": tid} for tid in to_clear]

    if UPDATE_VISIBLE_NAME and desired_visible_name and desired_visible_name != host.get("name"):
        params["name"] = desired_visible_name

    if UPDATE_TECHNICAL_HOSTNAME and desired_technical_host and desired_technical_host != host.get("host"):
        params["host"] = desired_technical_host

    if dry_run:
        print(f"[DRY] host.update params: {json.dumps(params, ensure_ascii=False)}")
        return

    api_call("host.update", params)


def create_host(visible_name: str, ip: str, groupids: List[str], templateids: List[str], dry_run: bool) -> None:
    tech = sanitize_technical_name(visible_name)
    params = {
        "host": tech,
        "name": visible_name,
        "interfaces": [{
            "type": 1,
            "main": 1,
            "useip": 1,
            "ip": ip,
            "dns": "",
            "port": AGENT_PORT
        }],
        "groups": [{"groupid": gid} for gid in groupids],
        "templates": [{"templateid": tid} for tid in templateids]
    }

    if dry_run:
        print(f"[DRY] host.create params: {json.dumps(params, ensure_ascii=False)}")
        return

    api_call("host.create", params)


def process_row(nome_visible: str, grupos_raw: Any, ip: str, templates_raw: Any, dry_run: bool) -> Tuple[str, str]:
    grupos = split_multi(grupos_raw)
    templates = split_multi(templates_raw)

    if not nome_visible or not ip:
        raise ValueError("Linha inválida: 'Nome' e 'IP' são obrigatórios.")
    if not grupos:
        raise ValueError("Linha inválida: pelo menos 1 'Grupo' é obrigatório.")
    if not templates:
        raise ValueError("Linha inválida: pelo menos 1 'Template' é obrigatório.")

    # Para hosts, usamos HOST GROUPS
    groupids = [ensure_hostgroup(g, dry_run=dry_run) for g in grupos]
    
    # Para templates, usamos a função ensure_template que já usa template groups
    templateids = [ensure_template(t, dry_run=dry_run) for t in templates]

    desired_tech = sanitize_technical_name(nome_visible)

    host = host_get_by_host_or_name(nome_visible)
    matched_by = "nome (host/name)"

    if not host:
        host = get_host_by_ip(ip)
        matched_by = "ip" if host else "novo"

    if not host:
        print(f"[CRIAR] {nome_visible} ({ip}) | Grupos={grupos} | Templates={templates}")
        create_host(nome_visible, ip, groupids, templateids, dry_run=dry_run)
        return ("WOULD_CREATE" if dry_run else "CREATED",
                f"Host não existia; {'seria criado' if dry_run else 'criado'} (name='{nome_visible}', host='{desired_tech}')")

    hostid = host["hostid"]
    print(f"[ATUALIZAR] {nome_visible} ({ip}) -> match por {matched_by} (hostid={hostid}, host='{host.get('host')}', name='{host.get('name')}')")

    ifaceid = pick_agent_interfaceid(host)
    update_interface_ip(hostid, ifaceid, ip, dry_run=dry_run)
    set_groups_and_templates(
        host,
        groupids,
        templateids,
        desired_visible_name=nome_visible,
        desired_technical_host=desired_tech,
        dry_run=dry_run
    )

    return ("WOULD_UPDATE" if dry_run else "UPDATED",
            f"Host existente (match por {matched_by}); {'seria atualizado' if dry_run else 'atualizado'} (hostid={hostid})")


def resolve_excel_path(excel_arg: str) -> str:
    if excel_arg:
        p = Path(excel_arg)
        if p.exists():
            return str(p)
    script_dir = Path(__file__).resolve().parent
    candidate = script_dir / (excel_arg or DEFAULT_EXCEL_NAME)
    if candidate.exists():
        return str(candidate)
    return excel_arg or DEFAULT_EXCEL_NAME


def build_report_path(report_csv: str, excel_path: str) -> str:
    if report_csv.strip():
        return report_csv.strip()
    excel_dir = Path(excel_path).resolve().parent
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return str(excel_dir / f"zabbix_import_report_{ts}.csv")


def write_report_csv(path: str, rows: List[Dict[str, Any]]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    fields = ["linha_excel", "nome", "ip", "grupos", "templates", "acao", "mensagem"]
    with p.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="", help="Caminho do Excel.")
    parser.add_argument("--report", default="", help="Caminho do CSV de relatório. Se omitido, salva ao lado do Excel.")
    parser.add_argument("--apply", action="store_true", help="Executa de verdade (DRY_RUN=False). Se omitido, apenas simula.")
    args = parser.parse_args()

    dry_run = not args.apply
    excel_path = resolve_excel_path(args.excel)
    report_path = build_report_path(args.report, excel_path)

    report_rows: List[Dict[str, Any]] = []

    print(f"[INFO] Zabbix Import Tool - Versão 7.4")
    print(f"[INFO] DRY_RUN={dry_run}")
    print(f"[INFO] Excel: {excel_path}")
    print(f"[INFO] Report CSV: {report_path}")
    print(f"[INFO] Template Group Padrão: {DEFAULT_TEMPLATE_GROUP}")

    wb = load_workbook(excel_path)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Aba '{SHEET_NAME}' não encontrada no arquivo {excel_path}. Abas: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    total = ok = fail = 0
    excel_row_num = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        excel_row_num += 1
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        total += 1
        nome, grupo, ip, template = (row + (None, None, None, None))[:4]
        nome_s = "" if nome is None else str(nome).strip()
        ip_s = "" if ip is None else str(ip).strip()

        entry = {
            "linha_excel": excel_row_num,
            "nome": nome_s,
            "ip": ip_s,
            "grupos": "" if grupo is None else str(grupo),
            "templates": "" if template is None else str(template),
            "acao": "",
            "mensagem": ""
        }

        try:
            print("=" * 90)
            print(f"Excel linha {excel_row_num}: Nome={nome} | Grupo={grupo} | IP={ip} | Template={template}")
            acao, detalhe = process_row(nome_s, grupo, ip_s, template, dry_run=dry_run)
            entry["acao"] = acao
            entry["mensagem"] = detalhe
            ok += 1
        except Exception as e:
            fail += 1
            entry["acao"] = "ERROR"
            entry["mensagem"] = str(e)
            print(f"[ERRO] {e}", file=sys.stderr)
            if not CONTINUE_ON_ERROR:
                report_rows.append(entry)
                write_report_csv(report_path, report_rows)
                print(f"[RELATÓRIO] CSV gerado: {report_path}")
                return 2

        report_rows.append(entry)

    write_report_csv(report_path, report_rows)
    print("=" * 90)
    print(f"Finalizado. Total processado={total} | OK={ok} | Falhas={fail} | DRY_RUN={dry_run}")
    print(f"[RELATÓRIO] CSV gerado: {report_path}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
